#! python3

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


"""Generates a multiplication table of X by X dimensions from user input."""

# Verifies user input by checking to see if they entered a positive integer. Loops until a valid input is received.
flag = False
while not flag:
    try:
        dimension = int(input("Please enter the dimension of the table: "))
        if dimension < 1:
            print("Must be an integer greater than 0.")
            continue
        else:
            print(f"Creating a {dimension}x{dimension} table...")
            flag = True
    except ValueError:
        print("Invalid input. Cannot be a letter, special character, negative number or floating point.")

mulTableWorkBook = openpyxl.Workbook()

mulTableSheet = mulTableWorkBook.active

mulTableSheet.title = "Multiplication"

# Gets the range of cells by using users input plus 1. The plus 1 is needed because we are not starting from the first
# row/column.
rowHeaders = mulTableSheet[f"A2:A{dimension + 1}"]

columnHeaders = mulTableSheet[f"B1:{get_column_letter(dimension + 1)}1"]

# Add styling for the headers by creating a Font object.

headerStyle = Font(size=16, bold=True)

# columnHeaders consists of a single tuple within another tuple: Each cell is in the same tuple. I can access each
# element by passing an index.
for number in range(dimension):
    for cell in columnHeaders:
        cell[number].value = number + 1
        cell[number].font = headerStyle

# rowHeaders consists of tuples within a tuple: Each cell is in its own tuple. The for loop returns a tuple, containing
# just one cell object.
for cell in rowHeaders:
    cell[0].value = cell[0].row - 1
    cell[0].font = headerStyle

# Select area of workbook that will contain the products of the numbers
productRange = mulTableSheet[f"B2:{get_column_letter(dimension + 1)}{dimension + 1}"]

# Iterates over each tuple, which is a row of cell objects. The numbers to be multiplied can be retrieved by the
# numerical value of the rows and columns subtracted by 1.
for group in productRange:
    for number in range(len(group)):
        group[number].value = (group[number].row - 1) * (group[number].column - 1)

mulTableSheet.freeze_panes = "B2"
# Saves workbook in current working directory.
mulTableWorkBook.save('multiplicationTable.xlsx')
